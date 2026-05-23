import pandas as pd
import openpyxl

# Files ka naam
input_file = "QUOTATION 26-27.xlsx"
output_file = "Master_Quotation_Summary.xlsx"

try:
    # Workbook load karna (data_only=True se formulas ki jagah direct values aayengi)
    wb = openpyxl.load_workbook(input_file, data_only=True)
    master_data = []
    sl_no = 1

    # Har sheet ke andar loop chalana
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 1. Quotation No. (A8) - "Quotation No.:" text ko clean karne ke liye
        raw_quotation = ws['A8'].value
        quotation_no = str(raw_quotation).replace("Quotation No.:", "").strip() if raw_quotation else ""
        
        # 2. Quotation Date (I8)
        quotation_date = ws['I8'].value
        
        # ⚠️ BOSS, YEH 3 DETAILS AAPNE MENTION NAHI KI THI, INHE APNE EXCEL KE HISAAB SE UPDATE KAR LIJIYEGA:
        customer_name = ws['A5'].value     # <--- Update this cell (e.g., 'B5')
        address = ws['A6'].value           # <--- Update this cell (e.g., 'B6')
        fabrication_no = ws['A7'].value    # <--- Update this cell (e.g., 'B7')

        # Parts ki list Row 16 se 21 tak
        for row in range(16, 22):
            part_no = ws.cell(row=row, column=2).value       # Column B
            part_desc = ws.cell(row=row, column=3).value     # Column C
            qty = ws.cell(row=row, column=4).value           # Column D
            rate = ws.cell(row=row, column=5).value          # Column E

            # Agar us row mein Part No. aur Description blank hai, toh wo skip ho jayegi
            if (part_no is None or str(part_no).strip() == "") and (part_desc is None or str(part_desc).strip() == ""):
                continue

            # Data ko list mein add karna
            master_data.append([
                sl_no, quotation_no, quotation_date, customer_name,
                address, fabrication_no, part_no, part_desc, qty, rate
            ])
            sl_no += 1

    # Dataframe banakar nayi Excel file mein save karna
    columns = [
        "Sl. No.", "Quotation No.", "Quotation date", "Customer Name", 
        "Address", "Fabrication No.", "Part No.", "Part Description", "Qty.", "Rate"
    ]
    df = pd.DataFrame(master_data, columns=columns)
    df.to_excel(output_file, index=False)
    
    print(f"Boss, file successfully ban gayi hai! Check: {output_file}")

except FileNotFoundError:
    print(f"Error: '{input_file}' nahi mili. Please dhyan rakhein ki script aur excel file ek hi folder mein hon.")
except Exception as e:
    print(f"Kuch error aa gaya: {e}")
